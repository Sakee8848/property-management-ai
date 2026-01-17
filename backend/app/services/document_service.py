"""
文档处理服务
"""
import os
from typing import BinaryIO, Dict, Optional
from pathlib import Path
import PyMuPDF  # fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image
import pytesseract
from loguru import logger

from app.core.config import settings


class DocumentProcessor:
    """文档处理类"""
    
    def __init__(self):
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    async def save_file(self, file: BinaryIO, filename: str, property_id: int) -> Dict:
        """
        保存上传的文件
        
        Args:
            file: 文件对象
            filename: 文件名
            property_id: 物业ID
        
        Returns:
            包含文件信息的字典
        """
        try:
            # 创建物业专属目录
            property_dir = self.upload_dir / str(property_id)
            property_dir.mkdir(parents=True, exist_ok=True)
            
            # 生成唯一文件名
            file_ext = Path(filename).suffix
            safe_filename = f"{os.urandom(16).hex()}{file_ext}"
            file_path = property_dir / safe_filename
            
            # 保存文件
            content = await file.read()
            with open(file_path, 'wb') as f:
                f.write(content)
            
            file_size = len(content)
            
            logger.info(f"文件保存成功: {file_path}")
            
            return {
                "file_name": filename,
                "file_path": str(file_path),
                "file_type": file_ext.lstrip('.'),
                "file_size": file_size,
            }
        
        except Exception as e:
            logger.error(f"保存文件错误: {str(e)}")
            raise
    
    def extract_text(self, file_path: str, file_type: str) -> str:
        """
        从文件中提取文本
        
        Args:
            file_path: 文件路径
            file_type: 文件类型
        
        Returns:
            提取的文本内容
        """
        try:
            if file_type == 'pdf':
                return self._extract_from_pdf(file_path)
            elif file_type in ['doc', 'docx']:
                return self._extract_from_docx(file_path)
            elif file_type in ['xls', 'xlsx']:
                return self._extract_from_excel(file_path)
            elif file_type == 'txt':
                return self._extract_from_txt(file_path)
            elif file_type in ['jpg', 'jpeg', 'png']:
                return self._extract_from_image(file_path)
            else:
                logger.warning(f"不支持的文件类型: {file_type}")
                return ""
        
        except Exception as e:
            logger.error(f"提取文本错误: {str(e)}")
            return ""
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """从PDF提取文本"""
        try:
            doc = PyMuPDF.open(file_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except Exception as e:
            logger.error(f"PDF提取错误: {str(e)}")
            return ""
    
    def _extract_from_docx(self, file_path: str) -> str:
        """从Word文档提取文本"""
        try:
            doc = DocxDocument(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            return text
        except Exception as e:
            logger.error(f"Word提取错误: {str(e)}")
            return ""
    
    def _extract_from_excel(self, file_path: str) -> str:
        """从Excel提取文本"""
        try:
            wb = load_workbook(file_path, read_only=True)
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
            wb.close()
            return text
        except Exception as e:
            logger.error(f"Excel提取错误: {str(e)}")
            return ""
    
    def _extract_from_txt(self, file_path: str) -> str:
        """从文本文件提取内容"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            with open(file_path, 'r', encoding='gbk') as f:
                return f.read()
        except Exception as e:
            logger.error(f"文本文件提取错误: {str(e)}")
            return ""
    
    def _extract_from_image(self, file_path: str) -> str:
        """从图片提取文本(OCR)"""
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image, lang='chi_sim')
            return text
        except Exception as e:
            logger.error(f"图片OCR错误: {str(e)}")
            return ""
    
    def delete_file(self, file_path: str):
        """删除文件"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"文件删除成功: {file_path}")
        except Exception as e:
            logger.error(f"删除文件错误: {str(e)}")
